"""
written by yan li
September , 2024

this function is used for modifing image name based on fileNamesCorrection.xlsx
in [ casia1groundtruth-master/CASIA 1.0 groundtruth ]
(download: https://github.com/CauchyComplete/casia1groundtruth)
"""

import openpyxl
import os


def load_image_rename_paths(filepath):
    workbook = openpyxl.load_workbook(filepath)
    # sheet = workbook.active       #获取当前活动表格
    sheet = workbook.worksheets[0]
    data_list = []

    for row in sheet.iter_rows(min_row=3,max_row=11,min_col=2,max_col=3,values_only=True):
        data_list.append(row)

    for row in sheet.iter_rows(min_row=14,max_row=16,min_col=2,max_col=3,values_only=True):
        data_list.append(row)

    for row in sheet.iter_rows(min_row=19,max_row=38,min_col=2,max_col=3,values_only=True):
        data_list.append(row)

    for row in sheet.iter_rows(min_row=41,max_row=41,min_col=2,max_col=3,values_only=True):
        data_list.append(row)

    print('Total number of images need to modify name is {}'.format(len(data_list)))
    return data_list


def modify_image(rename_images:list,img_dir):
    for src,tgt in rename_images:
        try:
            old_path = os.path.join(img_dir,src)
            new_path = os.path.join(img_dir,tgt)
            os.rename(old_path,new_path)
        except:
            print('Rename Error for name {}'.format(src))



if __name__ == '__main__':

    rename_fpth = r'E:\tamper_dataset\casia1groundtruth-master\CASIA 1.0 groundtruth\FileNameCorrection.xlsx'
    ren_img_list = load_image_rename_paths(rename_fpth)
    img_dir = r'E:\tamper_dataset\casia-dataset\CASIA1\Sp'
    modify_image(ren_img_list,img_dir)
    print(f'The images in the folder {img_dir} has been renamed ')